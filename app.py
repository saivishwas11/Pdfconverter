from flask import Flask, request, send_file, abort, render_template
import os
import tempfile
from werkzeug.utils import secure_filename
from docx2pdf import convert as docx_convert
import win32com.client as win32
import magic
import pythoncom
import uuid
import pandas as pd
from bs4 import BeautifulSoup
import markdown
import html2text
from io import StringIO
import time
import re
import webcolors
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side  
from openpyxl.utils import get_column_letter

app = Flask(__name__)

ALLOWED_EXTENSIONS = {'docx', 'xlsx', 'html'}
MIME_TYPES = {
    'docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'html': 'text/html'
}
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100 MB

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def validate_mime_type(filepath, ext):
    try:
        mime = magic.from_file(filepath, mime=True)
        expected_mime = MIME_TYPES.get(ext)
        return mime == expected_mime
    except Exception as e:
        print(f"Error validating MIME type for {filepath}: {e}")
        return False

def convert_excel_to_pdf(input_file, output_file):
    pythoncom.CoInitialize()
    excel = None
    wb = None
    try:
        excel = win32.DispatchEx('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(os.path.abspath(input_file))

        for sheet in wb.Sheets:
            page_setup = sheet.PageSetup
            page_setup.Zoom = False
            page_setup.FitToPagesWide = 1
            page_setup.FitToPagesTall = False
            page_setup.Orientation = 2
            page_setup.TopMargin = excel.InchesToPoints(0.25)
            page_setup.BottomMargin = excel.InchesToPoints(0)
            page_setup.LeftMargin = excel.InchesToPoints(0)
            page_setup.RightMargin = excel.InchesToPoints(0)
            page_setup.HeaderMargin = excel.InchesToPoints(0.25)
            page_setup.FooterMargin = excel.InchesToPoints(0.25)

        wb.ExportAsFixedFormat(0, os.path.abspath(output_file), 0)

    finally:
        if wb is not None:
            try:
                wb.Close(False)
            except Exception as e:
                print(f"Error closing workbook: {e}")
            del wb
        if excel is not None:
            try:
                excel.Quit()
            except Exception as e:
                print(f"Error quitting Excel: {e}")
            del excel
        pythoncom.CoUninitialize()

def html_color_to_openpyxl_argb(html_color):
    if not html_color:
        return None
    
    html_color = html_color.lower().strip()
    
    try:
        if html_color.startswith('#'):
            hex_val = html_color.lstrip('#')
        else:
            hex_val = webcolors.name_to_hex(html_color).lstrip('#')

        if len(hex_val) == 3:
            hex_val = "".join([c*2 for c in hex_val])
            
        return 'FF' + hex_val.upper()
    except ValueError:
        return None

def convert_to_excel(input_file, output_file):
    if input_file.endswith(('.docx', '.xlsx')):
        df = pd.read_excel(input_file, engine='openpyxl')
        df.to_excel(output_file, index=False)
        return

    elif input_file.endswith('.html'):
        with open(input_file, 'r', encoding='utf-8') as f:
            html_content = f.read()

        soup = BeautifulSoup(html_content, 'html.parser')
        tables = soup.find_all('table')

        if not tables:
            text = soup.get_text()
            df = pd.DataFrame({'Content': [text]})
            df.to_excel(output_file, index=False)
            return

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            workbook = writer.book
            if 'Sheet1' in workbook.sheetnames:
                 std = workbook['Sheet1']
                 workbook.remove(std)
            worksheet = workbook.create_sheet('Sheet1')

            thin_black_side = Side(style='thin', color='FF000000')
            default_border = Border(
                left=thin_black_side,
                right=thin_black_side,
                top=thin_black_side,
                bottom=thin_black_side
            )

            current_row_excel = 1
            for table in tables:
                rows = table.find_all('tr')
                for row in rows:
                    cells = row.find_all(['td', 'th'])
                    current_col_excel = 1
                    
                    for cell in cells:
                        text = cell.get_text(strip=True)
                        style = cell.get('style', '')
                        bg_color_html = cell.get('bgcolor')
                        font_color_html = None
                        text_align = 'general'

                        if style:
                            bg_match = re.search(r'background-color:\s*([^;]+)', style)
                            if bg_match:
                                bg_color_html = bg_match.group(1)
                            color_match = re.search(r'(?<!background-)color:\s*([^;]+)', style)
                            if color_match:
                                font_color_html = color_match.group(1)
                            align_match = re.search(r'text-align:\s*([^;]+)', style)
                            if align_match:
                                text_align = align_match.group(1).strip()

                        fill = None
                        font = Font()
                        alignment = Alignment(horizontal=text_align, vertical='center', wrap_text=True)

                        bg_color_argb = html_color_to_openpyxl_argb(bg_color_html)
                        if bg_color_argb:
                            fill = PatternFill(start_color=bg_color_argb, end_color=bg_color_argb, fill_type="solid")

                        font_color_argb = html_color_to_openpyxl_argb(font_color_html)
                        if font_color_argb:
                            font.color = font_color_argb
                        
                        if cell.find('b') or cell.name == 'th' or 'bold' in style:
                            font.bold = True

                        target_cell = worksheet.cell(row=current_row_excel, column=current_col_excel)
                        target_cell.value = text
                        target_cell.alignment = alignment
                        
                        if fill:
                            target_cell.fill = fill
                        if font.color or font.bold:
                            target_cell.font = font
                        
                        target_cell.border = default_border

                        colspan = int(cell.get('colspan', 1))
                        rowspan = int(cell.get('rowspan', 1))
                        if colspan > 1 or rowspan > 1:
                            worksheet.merge_cells(
                                start_row=current_row_excel,
                                start_column=current_col_excel,
                                end_row=current_row_excel + rowspan - 1,
                                end_column=current_col_excel + colspan - 1
                            )
                        
                        current_col_excel += colspan
                    
                    current_row_excel += 1

                current_row_excel += 1
            
            for col in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = min(adjusted_width, 60)

@app.route('/')
def index():
    return render_template('pdf.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    temp_output = None

    with tempfile.TemporaryDirectory() as tmpdirname:
        filepath = None
        output_file = None

        try:
            if 'file' not in request.files:
                abort(400, 'No file part in the request.')
            file = request.files['file']
            if file.filename == '':
                abort(400, 'No selected file.')
            if not allowed_file(file.filename):
                abort(400, f'Unsupported file type. Allowed types: {", ".join(ALLOWED_EXTENSIONS)}')

            output_format = request.form.get('output_format', 'pdf')
            if output_format not in ['pdf', 'excel']:
                abort(400, 'Invalid output format selected.')

            filename = secure_filename(file.filename)
            filepath = os.path.join(tmpdirname, filename)
            file.save(filepath)
            ext = filename.rsplit('.', 1)[1].lower()

            if not validate_mime_type(filepath, ext):
                abort(400, 'File type mismatch. Possible malicious or corrupted file.')

            output_extension = '.pdf' if output_format == 'pdf' else '.xlsx'
            output_file = os.path.join(tmpdirname, f'converted{output_extension}')

            if output_format == 'pdf':
                if ext == 'docx':
                    pythoncom.CoInitialize()
                    docx_convert(filepath, output_file)
                    pythoncom.CoUninitialize()
                elif ext == 'xlsx':
                    convert_excel_to_pdf(filepath, output_file)
                elif ext == 'html':
                    from weasyprint import HTML
                    HTML(string=open(filepath, 'r', encoding='utf-8').read()).write_pdf(output_file)
            else:
                convert_to_excel(filepath, output_file)

            temp_output = os.path.join(tempfile.gettempdir(), f'converted_{uuid.uuid4().hex}{output_extension}')
            with open(output_file, 'rb') as src, open(temp_output, 'wb') as dst:
                dst.write(src.read())

        except Exception as e:
            print(f"Unexpected error: {e}")
            abort(500, 'Internal server error.')

    return send_file(temp_output, as_attachment=True, download_name=f'converted{output_extension}')

if __name__ == '__main__':
    app.run(debug=False)
