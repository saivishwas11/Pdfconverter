from flask import Flask, request, send_file, abort, render_template
import os
import tempfile
from werkzeug.utils import secure_filename
from docx2pdf import convert as docx_convert
import win32com.client as win32
import magic
import pythoncom
import uuid
import pandas as pd
from bs4 import BeautifulSoup
import math
import re
import webcolors
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
import logging
import traceback
from zipfile import ZipFile

# Configure logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

app = Flask(__name__)

ALLOWED_EXTENSIONS = {'docx', 'xlsx', 'html'}
MIME_TYPES = {
    'docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'html': 'text/html'
}
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100 MB

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def validate_mime_type(filepath, ext):
    try:
        mime = magic.from_file(filepath, mime=True)
        expected_mime = MIME_TYPES.get(ext)
        logger.debug(f"File MIME type: {mime}, Expected MIME type: {expected_mime}")
        return mime == expected_mime
    except Exception as e:
        logger.error(f"Error validating MIME type for {filepath}: {e}")
        return False

def convert_to_excel(input_file, output_file):
    PIXELS_TO_EXCEL_UNITS = 8.43 

    with open(input_file, 'r', encoding='utf-8') as f:
        html_content = f.read()

    soup = BeautifulSoup(html_content, 'html.parser')
    tables = soup.find_all('table')

    if not tables:
        text = soup.get_text(separator='\n', strip=True)
        df = pd.DataFrame([line for line in text.split('\n') if line], columns=['Content'])
        df.to_excel(output_file, index=False)
        return

    workbook = Workbook()
    worksheet = workbook.active
    
    thin_black_side = Side(style='thin', color='FF000000')
    default_border = Border(left=thin_black_side, right=thin_black_side, top=thin_black_side, bottom=thin_black_side)

    master_layout_pixels = []
    max_cols = 0
    for table in tables:
        cols = table.find_all('col')
        if len(cols) > max_cols:
            max_cols = len(cols)
            master_layout_pixels = []
            for col in cols:
                style = col.get('style', '')
                match = re.search(r'width:\s*(\d+)', style)
                if match:
                    master_layout_pixels.append(int(match.group(1)))

    if not master_layout_pixels:
        logger.error("Could not determine a master layout from <colgroup> tags.")
        pd.read_html(html_content).to_excel(output_file, index=False)
        return

    master_layout_excel_units = [px / PIXELS_TO_EXCEL_UNITS for px in master_layout_pixels]
    for i, width in enumerate(master_layout_excel_units):
        worksheet.column_dimensions[get_column_letter(i + 1)].width = width

    current_row_excel = 1
    for table in tables:
        local_layout_pixels = []
        local_cols = table.find_all('col')
        if local_cols:
            for col in local_cols:
                style = col.get('style', '')
                match = re.search(r'width:\s*(\d+)', style)
                if match: local_layout_pixels.append(int(match.group(1)))

        rows = table.find_all('tr')
        for row in rows:
            cells = row.find_all(['td', 'th'])
            current_col_excel = 1
            
            for cell_idx, cell in enumerate(cells):
                text = cell.get_text(strip=True)
                style_str = cell.get('style', '') + row.get('style', '')
                
                bg_color_html = cell.get('bgcolor')
                if not bg_color_html:
                    bg_match = re.search(r'background-color:\s*([^;]+)', style_str)
                    if bg_match: bg_color_html = bg_match.group(1).strip()
                font_color_html = None
                color_match = re.search(r'(?<!background-)color:\s*([^;]+)', style_str)
                if color_match: font_color_html = color_match.group(1).strip()
                align_map = {'center': 'center', 'left': 'left', 'right': 'right', 'justify': 'justify'}
                text_align = 'general'
                align_match = re.search(r'text-align:\s*([^;]+)', style_str)
                if align_match: text_align = align_map.get(align_match.group(1).strip().lower(), 'general')
                is_bold = 'font-weight: bold' in style_str or cell.find('b') or cell.name == 'th'

                html_colspan = int(cell.get('colspan', 1))
                
                target_pixel_width = 0
                if local_layout_pixels and cell_idx < len(local_layout_pixels):
                    for i in range(html_colspan):
                        if (cell_idx + i) < len(local_layout_pixels):
                            target_pixel_width += local_layout_pixels[cell_idx + i]
                
                excel_colspan = 0
                covered_width = 0
                if target_pixel_width > 0:
                    start_master_col_idx = current_col_excel - 1
                    while covered_width < (target_pixel_width * 0.9) and (start_master_col_idx + excel_colspan) < len(master_layout_pixels):
                        covered_width += master_layout_pixels[start_master_col_idx + excel_colspan]
                        excel_colspan += 1
                excel_colspan = max(1, excel_colspan)

                alignment = Alignment(horizontal=text_align, vertical='center', wrap_text=True)
                font = Font(bold=bool(is_bold))
                fill = None
                bg_color_argb = html_color_to_openpyxl_argb(bg_color_html)
                if bg_color_argb:
                    try: fill = PatternFill(start_color=bg_color_argb, end_color=bg_color_argb, fill_type="solid")
                    except ValueError: fill = None
                font_color_argb = html_color_to_openpyxl_argb(font_color_html)
                if font_color_argb:
                    try: font.color = font_color_argb
                    except ValueError: pass

                target_cell = worksheet.cell(row=current_row_excel, column=current_col_excel)
                target_cell.value = text
                target_cell.alignment = alignment
                if fill: target_cell.fill = fill
                target_cell.font = font

                if excel_colspan > 1:
                    end_col = current_col_excel + excel_colspan - 1
                    worksheet.merge_cells(start_row=current_row_excel, start_column=current_col_excel, end_row=current_row_excel, end_column=end_col)
                    for r_offset in range(1):
                        for c_offset in range(excel_colspan):
                             worksheet.cell(row=current_row_excel + r_offset, column=current_col_excel + c_offset).border = default_border
                else:
                    target_cell.border = default_border
                
                current_col_excel += excel_colspan
            current_row_excel += 1
        current_row_excel += 1

    POINTS_PER_LINE = 15.0
    for row_index in range(1, worksheet.max_row + 1):
        max_lines_in_row = 1
        for cell in worksheet[row_index]:
            if not cell.value: continue
            
            effective_width_units = 0
            is_merged = False
            for merged_range in worksheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                        effective_width_units += worksheet.column_dimensions[get_column_letter(col_idx)].width
                    is_merged = True
                    break
            if not is_merged:
                effective_width_units = worksheet.column_dimensions[cell.column_letter].width

            text = str(cell.value)
            lines_from_newlines = text.count('\n') + 1
            lines_from_wrapping = 1
            if effective_width_units > 0:
                lines_from_wrapping = math.ceil(len(text) / (effective_width_units / 1.1))

            cell_lines = max(lines_from_newlines, lines_from_wrapping)
            if cell_lines > max_lines_in_row:
                max_lines_in_row = cell_lines
        
        worksheet.row_dimensions[row_index].height = max_lines_in_row * POINTS_PER_LINE

    workbook.save(output_file)
    
def html_color_to_openpyxl_argb(html_color):
    if not html_color:
        return None
    
    html_color = html_color.lower().strip()
    
    try:
        if html_color.startswith('#'):
            hex_val = html_color.lstrip('#')
        else:
            hex_val = webcolors.name_to_hex(html_color).lstrip('#')

        if len(hex_val) == 3:
            hex_val = "".join([c*2 for c in hex_val])
        
        if len(hex_val) == 6:
            return 'FF' + hex_val.upper()
        else:
            return None
            
    except ValueError:
        return None
        
@app.route('/')
def index():
    return render_template('pdf.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    temp_output = None
    output_extension = '.xlsx'

    try:
        with tempfile.TemporaryDirectory() as tmpdirname:
            output_files = []
            base_filenames = []
            if 'file' not in request.files:
                logger.error("No file part in the request")
                abort(400, 'No file part in the request.')
            
            files = request.files.getlist('file')
            if not files or all(f.filename == '' for f in files):
                logger.error("No selected file")
                abort(400, 'No selected file.')
            
            if any(not allowed_file(f.filename) for f in files):
                logger.error(f"Unsupported file type in uploaded files")
                abort(400, f'Unsupported file type. Allowed types: {", ".join(ALLOWED_EXTENSIONS)}')

            output_format = request.form.get('output_format', 'pdf')
            if output_format not in ['pdf', 'excel']:
                logger.error(f"Invalid output format: {output_format}")
                abort(400, 'Invalid output format selected.')

            for file in files:
                filename = secure_filename(file.filename)
                filepath = os.path.join(tmpdirname, filename)
                file.save(filepath)
                ext = filename.rsplit('.', 1)[1].lower()

                if not validate_mime_type(filepath, ext):
                    logger.error(f"File type mismatch for {filename}")
                    abort(400, f'File type mismatch for {filename}. Possible malicious or corrupted file.')

                base_filename = os.path.splitext(filename)[0]
                base_filenames.append(base_filename)
                output_extension = '.pdf' if output_format == 'pdf' else '.xlsx'
                output_file = os.path.join(tmpdirname, f'{base_filename}{output_extension}')

                try:
                    if output_format == 'pdf':
                        if ext == 'docx':
                            import pythoncom
                            pythoncom.CoInitialize()
                            try:
                                docx_convert(filepath, output_file)
                            finally:
                                pythoncom.CoUninitialize()
                        elif ext == 'xlsx':
                            convert_excel_to_pdf(filepath, output_file)
                        elif ext == 'html':
                            from weasyprint import HTML
                            HTML(string=open(filepath, 'r', encoding='utf-8').read()).write_pdf(output_file)
                    else:
                        convert_to_excel(filepath, output_file)
                    output_files.append(output_file)
                except Exception as e:
                    logger.error(f"Error during file conversion: {str(e)}")
                    logger.error(traceback.format_exc())
                    abort(500, f'Error during file conversion: {str(e)}')

            # Create a zip file containing all outputs
            zip_output = os.path.join(tempfile.gettempdir(), f'converted_{uuid.uuid4().hex}.zip')
            with ZipFile(zip_output, 'w') as zipf:
                for i, file_path in enumerate(output_files):
                    arcname = f'{base_filenames[i]}{output_extension}'
                    zipf.write(file_path, arcname=arcname)
            temp_output = zip_output

    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        logger.error(traceback.format_exc())
        abort(500, 'Internal server error.')

    if not temp_output or not os.path.exists(temp_output):
        logger.error("Output file was not created or found")
        abort(500, 'Failed to create the output file.')

    download_filename = f'converted_files.zip'
    return send_file(temp_output, as_attachment=True, download_name=download_filename)

if __name__ == '__main__':
    app.run(debug=True)